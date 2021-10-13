from datetime import datetime
import pandas as pd
import shutil
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from Radiation_Pixel import radiation_pixel

# Ctrl+Shift+E to Run All File
# Alt+Shift+E to run single value
# Ctrl+/ to comment line

""" Make Execution path as current working directory """
parent_path = "/home/kanav/Documents/Windy"
os.chdir(parent_path)
print("Current Path:", os.getcwd())

""" Today date and time """
current_time = datetime.now()
month = current_time.strftime("%B")
time = current_time.strftime("%H:%M")
date = current_time.strftime("%d_%m_%Y")
time_block = {"05:55": "31:37", "07:25": "37:43", "08:55": "43:49", "10:25": "49:55", "11:55": "55:61", "13:25": "61:67", "14:55": "67:73"}
block_id = time_block.get(time)

if pd.isna(block_id):
    time_block = {"05:56": "31:37", "07:26": "37:43", "08:56": "43:49", "10:26": "49:55", "11:56": "55:61",
                  "13:26": "61:67", "14:56": "67:73"}
    block_id = time_block.get(time)

print("Date: {}, Block_Id: {}".format(time, block_id))

Windy_Csv = pd.read_csv('Location.csv')  # reading the CSV file
zoom_level = [7]

""" Dynamic path for Screenshot """
date_path = str(
    current_time.year) + "/" + f"{current_time.month:02}" + "/" + f"{current_time.day:02}" + "/" + f"{current_time.hour:02}" + "/" + f"{current_time.minute:02}"

""" Remove existing Screenshots """
if os.path.exists(parent_path + "/Images/"):
    shutil.rmtree(parent_path + "/Images/")

""" Save or Read Pixel value in XLSX format"""
pixel_path = parent_path + "/Pixel/" + month
if not os.path.exists(pixel_path):
    os.makedirs(pixel_path)

pixel_file_name = pixel_path + "/" + "Pixel_" + str(date) + ".xlsx"

for i in range(len(Windy_Csv)):
    pixel_df = pd.DataFrame()
    farm_name = Windy_Csv.loc[i]['FARM_NAME']
    latitude = Windy_Csv.loc[i]['LATITUDE']
    longitude = Windy_Csv.loc[i]['LONGITUDE']
    iteration = i
    total_location = Windy_Csv.shape[0] - 1
    pixel_file_existence = os.path.exists(pixel_file_name)  # Check existence of Pixel File

    if pixel_file_existence:
        workbook = openpyxl.load_workbook(filename=pixel_file_name)
        sheet_names = workbook.sheetnames
        if farm_name in sheet_names:
            work_sheet = workbook[farm_name]  # if sheet name exists read the sheet
            work_sheet_data = work_sheet.values  # work sheet data
            col_name = next(work_sheet_data)[0:]  # get first line in file as header
            pixel_df = pd.DataFrame(work_sheet_data, columns=col_name)  # Get Sheet values in dataframe format
            pixel_df = pixel_df.iloc[:, 1:]
            workbook.remove(work_sheet)  # Remove sheet otherwise will create a duplicate
    else:
        workbook = openpyxl.Workbook()

    pixel_data = radiation_pixel(latitude, longitude, zoom_level, farm_name, parent_path, date_path,
                                 block_id, iteration, total_location)
    pixel_data = pd.DataFrame(pixel_data, index=[0])
    pixel_df = pixel_df.append(pixel_data)
    new_work_sheet = workbook.create_sheet(farm_name)  # create sheet at last position ,(farm_name,0) at first position
    if 'Sheet' in workbook.sheetnames:  # Creating new workbook create new work 'sheet'; remove it
        workbook.remove(workbook['Sheet'])

    for idx, val in enumerate(dataframe_to_rows(pixel_df, index=True, header=True)):
        if idx == 1:    # to remove null row between col_name and col_value
            continue
        new_work_sheet.append(val)

    workbook.save(filename=pixel_file_name)  # Save Workbook




""" Create Date Time & Block Id dataframe same as python"""
# time_stamp = pd.date_range(current_time.strftime("%Y-%m-%d"), periods=96, freq='15T')
# dataframe_time = pd.DataFrame({'DATETIME': time_stamp, 'BLOCK_ID': list(range(1, 97))})
# dataframe_time.to_csv("timedf.csv")


